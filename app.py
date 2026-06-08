"""
鸡蛋库存登记助手 — Flask 应用入口

- GET  /                     首页 (录入界面)
- GET  /history              历史记录页
- GET  /api/today            获取今天的最新记录 (自动加载)
- POST /api/submit           提交/更新录入数据 (支持 upsert)
- GET  /api/history          获取历史记录列表 (JSON)
- GET  /api/history/<id>     获取某条记录详情 (JSON)
- PUT  /api/history/<id>     更新某条记录 (批量编辑)
- DELETE /api/history/<id>   删除某条记录
"""

import os
import sqlite3
from datetime import date, datetime

from flask import Flask, g, jsonify, render_template, request

# ==========================================
#  App factory
# ==========================================

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-eggnum-2026")

# Database path: honour EGGS_DB_DIR env var, fallback to instance/
_db_dir = os.environ.get("EGGS_DB_DIR", app.instance_path)
app.config["DATABASE"] = os.path.join(_db_dir, "eggnum.db")

os.makedirs(_db_dir, exist_ok=True)

# ==========================================
#  Preset templates (from design.md §3.1)
# ==========================================

PRESET_TEMPLATES: dict[str, list[int]] = {
    "农家蛋":   [30, 15],
    "五谷蛋":   [30, 15, 10],
    "虫草蛋":   [30, 15, 10],
    "小花蛋":   [30, 20, 15],
    "五黑初生蛋": [20],
    "初生蛋":   [20],
}

DEFAULT_STORE_NAME = "鹏泰(大福店)"


# ==========================================
#  Helpers: build flat item list in template order
# ==========================================

def build_ordered_items(
    quantities_by_key: dict[str, int] | None = None,
) -> list[dict]:
    """
    Build a flat list of {category, spec, quantity} in preset template order.
    quantities_by_key: optional dict keyed by "category_spec" → quantity.
    """
    result = []
    for cat in PRESET_TEMPLATES:
        for sp in PRESET_TEMPLATES[cat]:
            key = f"{cat}_{sp}"
            qty = quantities_by_key.get(key, 0) if quantities_by_key else 0
            result.append({"category": cat, "spec": sp, "quantity": qty})
    return result


def _item_key(category: str, spec: int) -> str:
    return f"{category}_{spec}"


# ==========================================
#  Database helpers
# ==========================================


def get_db() -> sqlite3.Connection:
    """Get a database connection for the current request."""
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db


@app.teardown_appcontext
def close_db(exception: object) -> None:
    """Close the database connection at the end of a request."""
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    """Create tables if they don't exist."""
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS records (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            store_name  TEXT NOT NULL DEFAULT '鹏泰(大福店)',
            record_date DATE NOT NULL,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS record_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id   INTEGER NOT NULL,
            category    TEXT NOT NULL,
            spec        INTEGER NOT NULL,
            quantity    INTEGER NOT NULL DEFAULT 0,
            sort_order  INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (record_id) REFERENCES records(id)
        );

        CREATE INDEX IF NOT EXISTS idx_items_record
            ON record_items(record_id);
        CREATE INDEX IF NOT EXISTS idx_records_date
            ON records(record_date);

        CREATE TABLE IF NOT EXISTS reserve_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            category    TEXT NOT NULL,
            spec        INTEGER NOT NULL,
            quantity    INTEGER NOT NULL DEFAULT 0,
            updated_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(category, spec)
        );

        CREATE TABLE IF NOT EXISTS reserve_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            record_date DATE NOT NULL,
            category    TEXT NOT NULL,
            spec        INTEGER NOT NULL,
            delta       INTEGER NOT NULL,
            linked      INTEGER NOT NULL DEFAULT 1,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_reserve_log_date
            ON reserve_log(record_date);

        CREATE TABLE IF NOT EXISTS attendance (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            record_date DATE NOT NULL,
            time_start  TEXT NOT NULL,
            time_end    TEXT NOT NULL,
            hours       REAL NOT NULL DEFAULT 0,
            note        TEXT DEFAULT '',
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_attendance_date
            ON attendance(record_date);
    """)
    # Migration: add linked column to existing reserve_log
    try:
        db.execute("ALTER TABLE reserve_log ADD COLUMN linked INTEGER DEFAULT 1")
    except Exception:
        pass
    db.commit()


# ==========================================
#  Text generation (core logic)
# ==========================================


def generate_output_text(
    store_name: str,
    record_date_str: str,
    items: list[dict],
) -> str:
    """
    Generate formatted output text.

    Key rule: quantity == 0 → blank after colon (not "0").
    See design.md §5.2.
    """
    lines = [f"{store_name} {record_date_str}"]

    for item in items:
        category = item["category"]
        spec = item["spec"]
        qty = item.get("quantity", 0)

        # ★ Core judgment: qty == 0 → empty string
        qty_str = str(qty) if qty > 0 else ""

        lines.append(f"{category}{spec}枚:{qty_str}")

    return "\n".join(lines)


def format_date_cn(d: date) -> str:
    """Convert a date object to Chinese format like '6月4日'."""
    return f"{d.month}月{d.day}日"



def _load_items_for_record(record_id: int) -> list[dict]:
    """Load all items for a single record, ordered by sort_order."""
    db = get_db()
    rows = db.execute(
        """SELECT category, spec, quantity
           FROM record_items
           WHERE record_id = ?
           ORDER BY sort_order""",
        (record_id,),
    ).fetchall()
    return [dict(r) for r in rows]


# ==========================================
#  Routes — Pages
# ==========================================


@app.route("/")
def index():
    """Home page — data entry form."""
    today = date.today()
    return render_template(
        "index.html",
        store_name=DEFAULT_STORE_NAME,
        date_display=format_date_cn(today),
        templates=PRESET_TEMPLATES,
    )


@app.route("/history")
def history():
    """History page — list past records.  Optimized: single batch query for all items."""
    db = get_db()
    records = db.execute(
        """
        SELECT id, store_name, record_date, created_at
        FROM records
        ORDER BY record_date DESC, created_at DESC
        """
    ).fetchall()

    if not records:
        return render_template("history.html", date_groups=[])

    # ★ Optimisation: batch-load all items in ONE query (fixes N+1)
    record_ids = [r["id"] for r in records]
    placeholders = ",".join("?" for _ in record_ids)
    all_items = db.execute(
        f"""SELECT record_id, category, spec, quantity
            FROM record_items
            WHERE record_id IN ({placeholders})
            ORDER BY record_id, sort_order""",
        record_ids,
    ).fetchall()

    # Group items by record_id
    items_by_record: dict[int, list[dict]] = {}
    for item in all_items:
        rid = item["record_id"]
        if rid not in items_by_record:
            items_by_record[rid] = []
        items_by_record[rid].append(
            {"category": item["category"], "spec": item["spec"], "quantity": item["quantity"]}
        )

    # Build date groups
    date_groups: list[dict] = []
    seen_dates: set[str] = set()

    for row in records:
        d = row["record_date"]
        if d not in seen_dates:
            seen_dates.add(d)
            date_groups.append({"date_display": d, "records": []})

        items_list = items_by_record.get(row["id"], [])
        full_text = generate_output_text(row["store_name"], d, items_list)

        # Preview: first 3 lines
        preview_lines = full_text.split("\n")
        preview = "\n".join(preview_lines[:3])
        if len(preview_lines) > 3:
            preview += "\n..."

        date_groups[-1]["records"].append(
            {
                "id": row["id"],
                "store_name": row["store_name"],
                "record_date": d,
                "text_preview": preview,
                "item_count": len([it for it in items_list if it["quantity"] > 0]),
            }
        )

    return render_template("history.html", date_groups=date_groups)


# ==========================================
#  Routes — API
# ==========================================


# ── Reserve (库存留存) ──


@app.route("/api/reserve")
def api_reserve():
    """Get all reserve quantities (cumulative, cross-day)."""
    db = get_db()
    rows = db.execute(
        "SELECT category, spec, quantity FROM reserve_items ORDER BY category, spec"
    ).fetchall()
    items = [dict(r) for r in rows]
    return jsonify({"items": items})


@app.route("/api/reserve", methods=["POST"])
def api_reserve_update():
    """
    Update reserve quantity for a single spec and sync with today's report.

    Body: { category, spec, delta }
    - delta > 0: increase reserve, decrease today's report
    - delta < 0: decrease reserve, increase today's report
    """
    data = _parse_json_body()
    if not data:
        return jsonify({"success": False, "error": "无效数据"}), 400

    category = data.get("category", "")
    spec = data.get("spec", 0)
    delta = data.get("delta", 0)
    report_date = data.get("date", None)  # None = linkage OFF, skip report sync

    if delta == 0:
        return jsonify({"success": False, "error": "delta 不能为 0"}), 400

    db = get_db()

    # Ensure linked column exists (migration)
    try:
        db.execute("ALTER TABLE reserve_log ADD COLUMN linked INTEGER DEFAULT 1")
    except Exception:
        pass

    # Upsert reserve item
    existing = db.execute(
        "SELECT id, quantity FROM reserve_items WHERE category = ? AND spec = ?",
        (category, spec),
    ).fetchone()

    if existing:
        new_qty = existing["quantity"] + delta
        if new_qty < 0:
            return jsonify({"success": False, "error": "留存不足"}), 400
        db.execute(
            "UPDATE reserve_items SET quantity = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (new_qty, existing["id"]),
        )
    else:
        if delta < 0:
            return jsonify({"success": False, "error": "留存不足"}), 400
        new_qty = delta
        db.execute(
            "INSERT INTO reserve_items (category, spec, quantity) VALUES (?, ?, ?)",
            (category, spec, delta),
        )

    # Sync with report only if date provided (linkage ON)
    if report_date:
        report_row = db.execute(
            "SELECT id FROM records WHERE record_date = ? ORDER BY created_at DESC LIMIT 1",
            (report_date,),
        ).fetchone()

        if report_row:
            item_row = db.execute(
                "SELECT id, quantity FROM record_items WHERE record_id = ? AND category = ? AND spec = ?",
                (report_row["id"], category, spec),
            ).fetchone()
            if item_row:
                new_report_qty = item_row["quantity"] - delta
                if new_report_qty < 0:
                    new_report_qty = 0
                db.execute(
                    "UPDATE record_items SET quantity = ? WHERE id = ?",
                    (new_report_qty, item_row["id"]),
                )

    # Log the change (linked=1 if synced with report, 0 if standalone)
    log_date = report_date if report_date else date.today().isoformat()
    linked = 1 if report_date else 0
    db.execute(
        "INSERT INTO reserve_log (record_date, category, spec, delta, linked) VALUES (?, ?, ?, ?, ?)",
        (log_date, category, spec, delta, linked),
    )

    db.commit()
    return jsonify({"success": True, "quantity": new_qty})


# ── Attendance (考勤打卡) ──


@app.route("/api/attendance")
def api_attendance_list():
    """List attendance entries. Query: ?days=3 (default 3 days) or ?from=&to="""
    db = get_db()
    days = request.args.get("days", type=int)
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    if date_from and date_to:
        rows = db.execute(
            "SELECT * FROM attendance WHERE record_date BETWEEN ? AND ? ORDER BY record_date DESC, time_start",
            (date_from, date_to),
        ).fetchall()
    elif days:
        rows = db.execute(
            "SELECT * FROM attendance WHERE record_date >= date('now', ?) ORDER BY record_date DESC, time_start",
            (f"-{days} days",),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM attendance ORDER BY record_date DESC, time_start"
        ).fetchall()

    return jsonify({"entries": [dict(r) for r in rows]})


@app.route("/api/attendance", methods=["POST"])
def api_attendance_create():
    """Create an attendance entry. Body: {record_date, time_start, time_end, hours, note}"""
    data = _parse_json_body()
    if not data:
        return jsonify({"success": False, "error": "无效数据"}), 400

    record_date = data.get("record_date", date.today().isoformat())
    time_start = data.get("time_start", "")
    time_end = data.get("time_end", "")
    hours = float(data.get("hours", 0))
    note = data.get("note", "")

    if not time_start or not time_end:
        return jsonify({"success": False, "error": "请选择时间"}), 400

    db = get_db()
    cursor = db.execute(
        """INSERT INTO attendance (record_date, time_start, time_end, hours, note)
           VALUES (?, ?, ?, ?, ?)""",
        (record_date, time_start, time_end, hours, note),
    )
    db.commit()

    return jsonify({"success": True, "id": cursor.lastrowid})


@app.route("/api/attendance/<int:entry_id>", methods=["PUT", "DELETE"])
def api_attendance_modify(entry_id: int):
    """Update or delete an attendance entry."""
    db = get_db()
    row = db.execute("SELECT id FROM attendance WHERE id = ?", (entry_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "记录不存在"}), 404

    if request.method == "DELETE":
        db.execute("DELETE FROM attendance WHERE id = ?", (entry_id,))
        db.commit()
        return jsonify({"success": True})

    # PUT: update
    data = _parse_json_body()
    if not data:
        return jsonify({"success": False, "error": "无效数据"}), 400

    updates = []
    params = []
    for field in ["record_date", "time_start", "time_end", "hours", "note"]:
        if field in data:
            updates.append(f"{field} = ?")
            val = data[field]
            if field == "hours":
                val = float(val)
            params.append(val)

    if not updates:
        return jsonify({"success": False, "error": "无更新字段"}), 400

    params.append(entry_id)
    db.execute(
        f"UPDATE attendance SET {', '.join(updates)} WHERE id = ?", params
    )
    db.commit()

    return jsonify({"success": True})


@app.route("/attendance-history")
def attendance_history_page():
    """Full attendance history page."""
    return render_template("attendance_history.html")


@app.route("/api/attendance-history")
def api_attendance_history():
    """Get all attendance entries for the full history page."""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM attendance ORDER BY record_date DESC, time_start"
    ).fetchall()
    entries = [dict(r) for r in rows]
    # Group by date
    from collections import OrderedDict
    groups = OrderedDict()
    for e in entries:
        d = e["record_date"]
        if d not in groups:
            groups[d] = {"entries": [], "total": 0}
        groups[d]["entries"].append(e)
        groups[d]["total"] += e["hours"]

    result = []
    for d, g in groups.items():
        result.append({
            "date": d,
            "total": round(g["total"], 2),
            "entries": g["entries"],
        })
    return jsonify({"groups": result})


@app.route("/api/attendance/export")
def api_attendance_export():
    """Generate Excel report matching the template format."""
    import io as _io, os as _os

    try:
        import openpyxl as _xl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        return jsonify({"success": False, "error": "openpyxl 未安装"}), 500

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if not date_from or not date_to:
        return jsonify({"success": False, "error": "请指定起止日期"}), 400

    db = get_db()
    rows = db.execute(
        """SELECT * FROM attendance
           WHERE record_date BETWEEN ? AND ?
           ORDER BY record_date, time_start""",
        (date_from, date_to),
    ).fetchall()

    # Load template (preserve all formatting — just clear values)
    template_path = _os.path.join(_os.path.dirname(__file__), "考勤报表_2026_03_28_to_04_30.xlsx")
    has_template = _os.path.exists(template_path)

    if has_template:
        wb = _xl.load_workbook(template_path)
        ws = wb.active
        # Only clear values in data rows (2 to max_row), keep formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "考勤报表"
        for i, h in enumerate(["日期", "时间段", "时长", "备注"], 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, size=11)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 14

    # Group by date
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    for r in rows:
        groups[r["record_date"]].append(r)

    # Fill data — only set values, don't touch formatting
    row_idx = 2
    total_hours = 0.0

    for d in sorted(groups.keys()):
        seg_count = len(groups[d])
        time_ranges = ", ".join(f"{r['time_start']}-{r['time_end']}" for r in groups[d])
        day_hours = sum(r["hours"] for r in groups[d])
        notes = "、".join(r["note"] for r in groups[d] if r["note"])
        total_hours += day_hours

        c_date = ws.cell(row=row_idx, column=1); c_date.value = d
        c_time = ws.cell(row=row_idx, column=2); c_time.value = time_ranges
        c_hours = ws.cell(row=row_idx, column=3); c_hours.value = day_hours
        c_note = ws.cell(row=row_idx, column=4); c_note.value = notes

        # Alignment: date + hours = vertical center; time range = top + wrap
        v_center = Alignment(vertical="center")
        top_wrap = Alignment(vertical="top", wrap_text=True)
        c_date.alignment = v_center
        c_time.alignment = top_wrap
        c_hours.alignment = v_center
        c_note.alignment = v_center

        # Multi-segment: increase row height
        if seg_count > 1:
            ws.row_dimensions[row_idx].height = 15 * seg_count

        row_idx += 1

    # Total row — only if not using template (template already has formatted total row)
    # Write value into the last data+1 row, keeping any existing formatting
    total_cell_date = ws.cell(row=row_idx, column=1)
    total_cell_hours = ws.cell(row=row_idx, column=3)
    total_cell_date.value = "合计"
    total_cell_hours.value = round(total_hours, 2)

    # If no template, bold the total row
    if not has_template:
        total_cell_date.font = Font(bold=True, size=11)
        total_cell_hours.font = Font(bold=True, size=11)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file

    filename = f"考勤报表_{date_from}_to_{date_to}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/reserve-history")
def reserve_history_page():
    """扣留历史记录页面"""
    return render_template("reserve_history.html")


@app.route("/api/reserve-history", methods=["GET", "DELETE"])
def api_reserve_history():
    """Get or clear reserve change log."""
    db = get_db()

    # Migration: ensure linked column exists
    try:
        db.execute("ALTER TABLE reserve_log ADD COLUMN linked INTEGER DEFAULT 1")
    except Exception:
        pass

    if request.method == "DELETE":
        dates = request.args.get("dates", "")
        if dates:
            date_list = [d.strip() for d in dates.split(",") if d.strip()]
            placeholders = ",".join("?" for _ in date_list)
            db.execute(
                f"DELETE FROM reserve_log WHERE record_date IN ({placeholders})",
                date_list,
            )
        else:
            db.execute("DELETE FROM reserve_log")
        db.commit()
        return jsonify({"success": True})

    rows = db.execute(
        """SELECT record_date, category, spec, delta, created_at
           FROM reserve_log ORDER BY record_date DESC, created_at DESC"""
    ).fetchall()

    from collections import OrderedDict
    groups = OrderedDict()
    for r in rows:
        d = r["record_date"]
        if d not in groups:
            groups[d] = {"date": d, "items": [], "total_delta": 0}
        groups[d]["items"].append({
            "category": r["category"],
            "spec": r["spec"],
            "delta": r["delta"],
            "created_at": r["created_at"],
            "linked": bool(r["linked"]) if "linked" in r.keys() else True,
        })
        groups[d]["total_delta"] += r["delta"]

    return jsonify({"groups": list(groups.values())})


@app.route("/api/reserve/log-event", methods=["POST"])
def api_reserve_log_event():
    """Log a system event (linkage toggle) to reserve_log."""
    data = _parse_json_body()
    if not data: return jsonify({"success": False}), 400
    db = get_db()
    try: db.execute("ALTER TABLE reserve_log ADD COLUMN linked INTEGER DEFAULT 1")
    except: pass
    db.execute(
        "INSERT INTO reserve_log (record_date, category, spec, delta, linked) VALUES (?,?,?,?,?)",
        (data.get("record_date", date.today().isoformat()), data.get("category","__link__"), data.get("spec",0), data.get("delta",0), 1)
    )
    db.commit()
    return jsonify({"success": True})


@app.route("/api/reserve/history/log", methods=["POST"])
def api_reserve_log_event():
    """Log a system event to reserve_log (e.g. linkage change)."""
    data = _parse_json_body()
    if not data:
        return jsonify({"success": False}), 400
    db = get_db()
    db.execute(
        "INSERT INTO reserve_log (record_date, category, spec, delta) VALUES (?, ?, ?, ?)",
        (data.get("record_date", date.today().isoformat()),
         data.get("category", "__system__"),
         data.get("spec", 0),
         data.get("delta", 0)),
    )
    db.commit()
    return jsonify({"success": True})


@app.route("/api/debug")
def api_debug():
    """Show ALL records and their items — for troubleshooting."""
    db = get_db()
    records = db.execute(
        "SELECT id, store_name, record_date, created_at FROM records ORDER BY record_date DESC, created_at DESC"
    ).fetchall()
    result = []
    for r in records:
        items = db.execute(
            "SELECT category, spec, quantity FROM record_items WHERE record_id = ? ORDER BY sort_order",
            (r["id"],),
        ).fetchall()
        qty_sum = sum(it["quantity"] for it in items)
        result.append({
            "id": r["id"],
            "record_date": r["record_date"],
            "created_at": r["created_at"],
            "total_qty": qty_sum,
            "items": [dict(it) for it in items if it["quantity"] > 0],
        })
    return jsonify({"record_count": len(result), "records": result})


@app.route("/api/today")
def api_today():
    """
    Get today's latest record for auto-loading on the index page.
    Accepts ?date=YYYY-MM-DD from client to avoid server timezone mismatch.
    Returns {found: true, record_id, items: [...]} or {found: false}.
    """
    today_str = request.args.get("date", date.today().isoformat())
    db = get_db()

    row = db.execute(
        """
        SELECT id, store_name, record_date
        FROM records
        WHERE record_date = ?
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (today_str,),
    ).fetchone()

    if not row:
        resp = jsonify({"found": False})
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return resp

    items = _load_items_for_record(row["id"])

    resp = jsonify(
        {
            "found": True,
            "record_id": row["id"],
            "store_name": row["store_name"],
            "record_date": row["record_date"],
            "items": items,
        }
    )
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


def _parse_json_body():
    """
    Parse JSON request body regardless of Content-Type header.
    sendBeacon sends text/plain, not application/json, so get_json() fails.
    """
    import json as _json

    data = request.get_json(silent=True)
    if data is not None:
        return data
    # Fallback: try parsing raw body as JSON
    raw = request.get_data(as_text=True)
    if raw:
        try:
            return _json.loads(raw)
        except _json.JSONDecodeError:
            pass
    return None


@app.route("/api/submit", methods=["POST"])
def api_submit():
    """
    Submit or update a record.

    Body:
    {
        "store_name": "...",
        "record_date": "YYYY-MM-DD",
        "items": [{category, spec, quantity}, ...],
        "record_id": null | int   // if provided, UPDATE existing; else INSERT
    }
    """
    data = _parse_json_body()
    if not data:
        return jsonify({"success": False, "error": "无效的请求数据"}), 400

    store_name = data.get("store_name", DEFAULT_STORE_NAME)
    record_date_str = data.get("record_date", "")
    items_in = data.get("items", [])
    record_id = data.get("record_id")
    merge_mode = data.get("merge", False)  # True → update only given items

    if not items_in:
        return jsonify({"success": False, "error": "没有提交任何数据"}), 400

    # Parse date
    try:
        record_date = datetime.strptime(record_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        record_date = date.today()

    # Build quantity lookup from submitted items
    qty_map: dict[str, int] = {}
    for item in items_in:
        key = _item_key(item.get("category", ""), item.get("spec", 0))
        qty = item.get("quantity", 0)
        qty = max(0, min(999, int(qty) if qty else 0))
        qty_map[key] = qty

    db = get_db()

    # ── Multi-device: always upsert by DATE. Keep only ONE record per date ──
    same_date = db.execute(
        "SELECT id FROM records WHERE record_date = ? ORDER BY created_at DESC",
        (record_date.isoformat(),),
    ).fetchall()

    if same_date:
        used_id = same_date[0]["id"]
        db.execute(
            "UPDATE records SET store_name = ?, record_date = ? WHERE id = ?",
            (store_name, record_date.isoformat(), used_id),
        )
        if not merge_mode:
            # Full replace: delete all items, re-insert
            db.execute("DELETE FROM record_items WHERE record_id = ?", (used_id,))
        # Clean up duplicates
        for dup in same_date[1:]:
            db.execute("DELETE FROM record_items WHERE record_id = ?", (dup["id"],))
            db.execute("DELETE FROM records WHERE id = ?", (dup["id"],))
    else:
        cursor = db.execute(
            "INSERT INTO records (store_name, record_date) VALUES (?, ?)",
            (store_name, record_date.isoformat()),
        )
        used_id = cursor.lastrowid

    # Upsert items
    if merge_mode and same_date:
        # Only update/insert the changed items
        for item in items_in:
            cat = item.get("category", "")
            sp = item.get("spec", 0)
            qty = max(0, min(999, int(item.get("quantity", 0))))
            existing_item = db.execute(
                "SELECT id FROM record_items WHERE record_id = ? AND category = ? AND spec = ?",
                (used_id, cat, sp),
            ).fetchone()
            if existing_item:
                db.execute(
                    "UPDATE record_items SET quantity = ? WHERE id = ?",
                    (qty, existing_item["id"]),
                )
            else:
                # Find max sort_order for this record
                max_sort = db.execute(
                    "SELECT COALESCE(MAX(sort_order), -1) FROM record_items WHERE record_id = ?",
                    (used_id,),
                ).fetchone()[0]
                db.execute(
                    "INSERT INTO record_items (record_id, category, spec, quantity, sort_order) VALUES (?, ?, ?, ?, ?)",
                    (used_id, cat, sp, qty, max_sort + 1),
                )
    else:
        # Full replace: insert all items in template order
        sort_idx = 0
        for cat in PRESET_TEMPLATES:
            for sp in PRESET_TEMPLATES[cat]:
                key = _item_key(cat, sp)
                qty = qty_map.get(key, 0)
                db.execute(
                    """INSERT INTO record_items
                       (record_id, category, spec, quantity, sort_order)
                       VALUES (?, ?, ?, ?, ?)""",
                    (used_id, cat, sp, qty, sort_idx),
                )
                sort_idx += 1

    db.commit()

    # Generate output text
    ordered_items = build_ordered_items(qty_map)
    text = generate_output_text(store_name, record_date.isoformat(), ordered_items)

    return jsonify(
        {
            "success": True,
            "text": text,
            "record_id": used_id,
            "is_update": bool(same_date),
        }
    )


@app.route("/api/history")
def api_history():
    """Get all history records as JSON (optimised batch query)."""
    db = get_db()
    records = db.execute(
        """
        SELECT id, store_name, record_date, created_at
        FROM records
        ORDER BY record_date DESC, created_at DESC
        """
    ).fetchall()

    if not records:
        return jsonify([])

    # Batch-load all items in one query
    record_ids = [r["id"] for r in records]
    placeholders = ",".join("?" for _ in record_ids)
    all_items = db.execute(
        f"""SELECT record_id, category, spec, quantity
            FROM record_items
            WHERE record_id IN ({placeholders})
            ORDER BY record_id, sort_order""",
        record_ids,
    ).fetchall()

    items_by_record: dict[int, list[dict]] = {}
    for item in all_items:
        rid = item["record_id"]
        if rid not in items_by_record:
            items_by_record[rid] = []
        items_by_record[rid].append(
            {"category": item["category"], "spec": item["spec"], "quantity": item["quantity"]}
        )

    result = []
    for row in records:
        items_list = items_by_record.get(row["id"], [])
        result.append(
            {
                "id": row["id"],
                "store_name": row["store_name"],
                "record_date": row["record_date"],
                "created_at": row["created_at"],
                "items": items_list,
                "text": generate_output_text(row["store_name"], row["record_date"], items_list),
            }
        )

    return jsonify(result)


@app.route("/api/history/<int:record_id>", methods=["GET", "PUT", "DELETE"])
def api_history_detail(record_id: int):
    """Get, update, or delete a single history record."""

    db = get_db()

    # --- GET: return detail ---
    if request.method == "GET":
        row = db.execute(
            "SELECT id, store_name, record_date, created_at FROM records WHERE id = ?",
            (record_id,),
        ).fetchone()

        if not row:
            return jsonify({"success": False, "error": "记录不存在"}), 404

        items = _load_items_for_record(record_id)

        return jsonify(
            {
                "success": True,
                "id": row["id"],
                "store_name": row["store_name"],
                "record_date": row["record_date"],
                "created_at": row["created_at"],
                "items": items,
                "text": generate_output_text(row["store_name"], row["record_date"], items),
            }
        )

    # --- DELETE: remove record and its items ---
    if request.method == "DELETE":
        row = db.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "记录不存在"}), 404

        db.execute("DELETE FROM record_items WHERE record_id = ?", (record_id,))
        db.execute("DELETE FROM records WHERE id = ?", (record_id,))
        db.commit()

        return jsonify({"success": True})

    # --- PUT: update record (batch edit) ---
    if request.method == "PUT":
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"success": False, "error": "无效的请求数据"}), 400

        row = db.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "记录不存在"}), 404

        # Update store_name if provided
        if "store_name" in data:
            db.execute(
                "UPDATE records SET store_name = ? WHERE id = ?",
                (data["store_name"], record_id),
            )

        # Update record_date if provided
        if "record_date" in data:
            db.execute(
                "UPDATE records SET record_date = ? WHERE id = ?",
                (data["record_date"], record_id),
            )

        # Update items if provided
        if "items" in data:
            for item in data["items"]:
                db.execute(
                    """UPDATE record_items
                       SET quantity = ?
                       WHERE record_id = ? AND category = ? AND spec = ?""",
                    (
                        max(0, min(999, int(item.get("quantity", 0)))),
                        record_id,
                        item["category"],
                        item["spec"],
                    ),
                )

        db.commit()

        # Return updated record
        items = _load_items_for_record(record_id)
        row2 = db.execute(
            "SELECT id, store_name, record_date, created_at FROM records WHERE id = ?",
            (record_id,),
        ).fetchone()

        return jsonify(
            {
                "success": True,
                "id": row2["id"],
                "store_name": row2["store_name"],
                "record_date": row2["record_date"],
                "items": items,
                "text": generate_output_text(row2["store_name"], row2["record_date"], items),
            }
        )

    # Should not reach here
    return jsonify({"success": False, "error": "不支持的请求方法"}), 405


@app.route("/api/history/<int:record_id>/text")
def api_history_text(record_id: int):
    """Re-generate text for a specific record (for one-tap copy)."""
    db = get_db()
    row = db.execute(
        "SELECT store_name, record_date FROM records WHERE id = ?",
        (record_id,),
    ).fetchone()

    if not row:
        return jsonify({"success": False, "error": "记录不存在"}), 404

    items = _load_items_for_record(record_id)
    text = generate_output_text(row["store_name"], row["record_date"], items)

    return jsonify({"success": True, "text": text})


# ==========================================
#  CLI init-db command
# ==========================================


@app.cli.command("init-db")
def init_db_command():
    """Create database tables."""
    with app.app_context():
        init_db()
    print("Database initialized.")


# ==========================================
#  Main
# ==========================================

# ── Auto-init DB on first request (Docker-friendly) ──
_init_done = False


@app.before_request
def _auto_init():
    global _init_done
    if not _init_done:
        _init_done = True
        init_db()


if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
